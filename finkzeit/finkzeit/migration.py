# -*- coding: utf-8 -*-
# Copyright (c) 2018-2019, Fink Zeitsysteme/libracore and contributors
# For license information, please see license.txt
#
# Run with
#   bench execute finkzeit.finkzeit.migration.import_customers --kwargs "{'filename': '/mnt/share/testing.xlsx'}"

# imports
from __future__ import unicode_literals
import frappe
from frappe import _
import csv
import codecs
from datetime import datetime
from openpyxl import load_workbook
from finkzeit.finkzeit.zsw import create_update_customer

# column allocation
FIRST_DATA_ROW = 4              # first row containing data (after headers)
CUSTOMER_ID = 1                 # K-##### (column B)
CUSTOMER_NAME = 2               # customer name (column C)
ADR_LINE_1 = 5                  # first address line (clumn E)
ADR_LINE_2 = 4                  # additional address text (column E)
ADR_PLZ = 6                     # pincode for address
ADR_CITY = 7                    # address city
ADR_COUNTRY = 8                 # country (code, e.g. AT)
CUSTOMER_KST = 9                # customer cost center (999: FZV, ???: FZT, ???: FZW)
CUSTOMER_DESCRIPTION = 46       # long description
CUSTOMER_CONDITIONS = 16        # payment conditions (TODO: clean up)
CUSTOMER_VAT_REGION = 15        # Steuerregion (1, ...)
CUSTOMER_CURRENCY = 14          # currency
BRIEFCODE = 12                  # letter salutation (code 1301, ...)
PHONE = 28                      # phone
FAX = 29                        # fax
HOMEPAGE = 37                   # homepage
EMAIL = 39                      # email
CUSTOMER_LSV = 40               # enable LSV
CUSTOMER_LSV_CODE = 41          # LSV code
CUSTOMER_LSV_DATE = 42          # LSV date (something like 46721)
CUSTOMER_IBAN = 43              # IBAN
CUSTOMER_BIC = 44               # BIC
CUSTOMER_TAX_ID = 45            # UST ID
CUSTOMER_INVOICE_SEND = 47      # type of invoice sending
WARTUNGSKUNDE = 49              # Wartungskunde
REFERENZKUNDE = 54              # Referenzkunde

def import_customers(filename, force_update=False):
    if force_update == "True" or force_update == 1:
        force_update = True

    # open workbook
    print("Loading {0}...".format(filename))
    workbook = load_workbook(filename)
    worksheet = workbook["Tabelle1"]
    for row in worksheet.iter_rows('A{}:BH{}'.format(FIRST_DATA_ROW, worksheet.max_row)):
        # loop through all customers
        #print(row)
        cells = row
        print("cells: {0}".format(len(cells)))
        if len(cells) >= 23:
            # check if customer exists by ID
            #print(cells[CUSTOMER_ID].value)
            matches_by_id = frappe.get_all("Customer", filters={'name': cells[CUSTOMER_ID].value}, fields=['name'])
            print("Customer: {0}".format(cells[CUSTOMER_ID].value))
            if matches_by_id:
                # found customer, update
                print("updating...")
                update_customer(matches_by_id[0]['name'], cells)
            else:
                # no match found by ID, check name with 0 (ID not set)
                print("creating...")
                create_customer(cells)
    return

def get_country_from_code(code):
    if not code or code == "":
        return "Schweiz"
    else:
        countries = frappe.get_all('Country', filters={'code':code.lower()}, fields=['name'])
        if countries:
            return countries[0]['name']
        else:
            return "Schweiz"

def get_kst_from_code(code):
    if str(code) == "050":
        return "FZW - FZAT"
    elif str(code) == "060":
        return "FZT - FZAT"
    else:
        return "FZV - FZAT"

def get_steuerregion_from_code(code):
    if (str(code) == "0" or str(code) == "C"):
        # CH/LI
        return "DRL"
    elif str(code) == "1":
        # AT
        return "AT"
    elif (str(code) == "2" or str(code) == "3" or str(code) == "A"):
        # DE, IT, HU, ...
        return "EU"                
    else:
        return "AT"

def get_date_from_excel(excel_date):
    if excel_date:
        dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + excel_date - 2)
        return dt
    else:
        return None
    
def create_customer(cells):
    # create record
    cus = frappe.get_doc(
        {
            "doctype":"Customer", 
            "name": cells[CUSTOMER_ID].value,
            "naming_series": "K-.#####",
            "customer_name": cells[CUSTOMER_NAME].value,
            "customer_type": "Company",
            "customer_group": "All Customer Groups",
            "territory": "All Territories",
            "description": cells[CUSTOMER_DESCRIPTION].value,
            "payment_terms": cells[CUSTOMER_CONDITIONS].value,
            "kostenstelle": get_kst_from_code(cells[CUSTOMER_KST].value),
            "steuerregion": get_steuerregion_from_code(cells[CUSTOMER_VAT_REGION].value),
            "default_currency": cells[CUSTOMER_CURRENCY].value,
            "website": cells[HOMEPAGE].value,
            "enable_lsv": cells[CUSTOMER_LSV].value,
            "lsv_code": cells[CUSTOMER_LSV_CODE].value,
            "lsv_date": get_date_from_excel(cells[CUSTOMER_LSV_DATE].value),
            "iban": cells[CUSTOMER_IBAN].value,
            "bic": cells[CUSTOMER_BIC].value,
            "tax_id": cells[CUSTOMER_TAX_ID].value,
        })
    try:
        new_customer = cus.insert()
        new_customer.name = cells[CUSTOMER_ID].value
        new_customer.save()
    except Exception as e:
        print(_("Insert customer failed"), _("Insert failed for customer {0}: {1}").format(
            cells[CUSTOMER_ID].value, e))
    else:
        #create_contact(cells, new_customer.name)
        create_address(cells, new_customer.name)
    # write changes to db
    frappe.db.commit()
    return

def create_contact(cells, customer):
    try:
        fullname = get_full_name(cells)
        if cells[VNAME].value == "":
            first_name = "-"
        else:
            first_name = cells[VNAME].value
        con = frappe.get_doc(
            {
                "doctype":"Contact", 
                "name": "{0} ({1})".format(fullname, customer),
                "first_name": get_first_name(cells),
                "last_name": cells[NNAME].value,
                "email_id": cells[EMAILADR].value,
                "salutation": cells[ANRED].value,
                "letter_salutation": cells[BRANRED].value,
                "fax": cells[TELEF].value,
                "phone": cells[TELEP].value,
                "mobile_no": cells[NATEL].value,
                "links": [
                    {
                        "link_doctype": "Customer",
                        "link_name": customer
                    }
                ]
            })
        new_contact = con.insert()
        return new_contact
    except Exception as e:
        print(_("Insert contact failed"), _("Insert failed for contact {0} {1} ({2}): {3}").format(
            cells[VNAME].value, cells[NNAME].value, cells[CUSTOMER_ID].value, e))
        return None

def create_address(cells, customer):
    try:
        adr = frappe.get_doc(
            {
                "doctype":"Address", 
                "name": "{0} ({1})".format(cells[CUSTOMER_NAME].value, customer),
                "address_title": "{0} ({1})".format(cells[CUSTOMER_NAME].value, customer),
                "address_line1": cells[ADR_LINE_1].value,
                "address_line2": cells[ADR_LINE_2].value,
                "city": cells[ADR_CITY].value,
                "pincode": cells[ADR_PLZ].value,
                "phone": cells[PHONE].value,
                "fax": cells[FAX].value,
                "email_id": cells[EMAIL].value,
                "is_primary_address": 1,
                "is_shipping_address": 1,
                "country": get_country_from_code(cells[ADR_COUNTRY].value),
                "links": [
                    {
                        "link_doctype": "Customer",
                        "link_name": customer
                    }
                ]
            })
        new_adr = adr.insert()
        return new_adr
    except Exception as e:
        print(_("Insert address failed"), _("Insert failed for address {0}: {1}").format(
            cells[CUSTOMER_ID].value, e))
        return None
    
def update_customer(name, cells):
    # get customer record
    cus = frappe.get_doc("Customer", name)
    cus.customer_name = cells[CUSTOMER_NAME].value
    #cus.customer_type = "Company"
    #cus.customer_group = "All Customer Groups"
    #cus.territory = "All Territories"
    description = cells[CUSTOMER_DESCRIPTION].value or ""
    cus.customer_details = description.replace("_x000D_", "")
    cus.payment_terms = cells[CUSTOMER_CONDITIONS].value
    cus.kostenstelle = get_kst_from_code(cells[CUSTOMER_KST].value)
    cus.steuerregion = get_steuerregion_from_code(cells[CUSTOMER_VAT_REGION].value)
    cus.default_currency = cells[CUSTOMER_CURRENCY].value
    cus.website = cells[HOMEPAGE].value
    cus.enable_lsv = cells[CUSTOMER_LSV].value
    cus.lsv_code = cells[CUSTOMER_LSV_CODE].value
    cus.lsv_date = get_date_from_excel(cells[CUSTOMER_LSV_DATE].value)
    cus.iban = cells[CUSTOMER_IBAN].value
    cus.bic = cells[CUSTOMER_BIC].value
    cus.tax_id = cells[CUSTOMER_TAX_ID].value
    cus.rechnungszustellung = cells[CUSTOMER_INVOICE_SEND].value
    cus.referenzkunde = cells[REFERENZKUNDE].value
    cus.wartungskunde = cells[WARTUNGSKUNDE].value
    try:
        cus.save()
    except Exception as e:
        print(_("Update customer failed"), _("Update failed for customer {0}: {1}").format(
            cells[CUSTOMER_ID].value, e))
    else:
        #con_id = frappe.get_all("Dynamic Link", 
        #    filters={'link_doctype': 'Customer', 'link_name': cus.name, 'parenttype': 'Contact'},
        #    fields=['parent'])
        #if con_id:
        #    # update contact
        #    con = frappe.get_doc("Contact", con_id[0]['parent'])
        #    con.first_name = get_first_name(cells)
        #    con.last_name = cells[NNAME].value or ''
        #    con.email_id = cells[EMAILADR].value or ''
        #    con.salutation = cells[ANRED].value or ''
        #    con.letter_salutation = cells[BRANRED].value or ''
        #    con.fax = cells[TELEF].value or ''
        #    con.phone = cells[TELEP].value or ''
        #    try:
        #        con.save()
        #    except Exception as e:
        #        print(_("Update contact failed"), _("Update failed for contact {0}: {1}").format(
        #            cells[CUSTOMER_ID].value, e))
        #else:
        #    # no contact available, create
        #    create_contact(cells, cus.name)
        adr_id = frappe.get_all("Dynamic Link", 
                filters={'link_doctype': 'Customer', 'link_name': cus.name, 'parenttype': 'Address'},
                fields=['parent'])
        if adr_id:
            adr = frappe.get_doc("Address", adr_id[0]['parent'])
            adr.address_title = "{0} ({1})".format(cells[CUSTOMER_NAME].value, name)
            adr.address_line1 = cells[ADR_LINE_1].value
            adr.address_line2 = cells[ADR_LINE_2].value
            adr.city = cells[ADR_CITY].value
            adr.pincode = cells[ADR_PLZ].value
            adr.phone = cells[PHONE].value
            adr.fax = cells[FAX].value
            adr.email_id = cells[EMAIL].value
            adr.is_primary_address = 1
            adr.is_shipping_address = 1
            adr.country = get_country_from_code(cells[ADR_COUNTRY].value)
            try:
                adr.save()
            except Exception as e:
                print(_("Update address failed"), _("Update address for contact {0}: {1}").format(
                    cells[CUSTOMER_ID].value, e))
        else:
            # address not found, create
            create_address(cells, cus.name)
    # write changes to db
    frappe.db.commit()
    return

# this function loops through all addresses and computes the customer matchcode
def get_matchcode():
    customers = frappe.get_all('Customer', filters=None, fields=['name'])
    for customer in customers:
        adr_id = frappe.get_all("Dynamic Link", 
                filters={'link_doctype': 'Customer', 'link_name': customer['name'], 'parenttype': 'Address'},
                fields=['parent'])
        if adr_id:
            adr = frappe.get_doc("Address", adr_id[0]['parent'])
            cus = frappe.get_doc("Customer", customer['name'])
            print("Matching {0}".format(customer.name))
            cus.rechnungsadresse = "{plz} {city}, {adr}".format(plz=adr.pincode, city=adr.city, adr=adr.address_line1)
            cus.save()
    return

# load taxes from customer record
def licence_add_taxes():
    licences = frappe.get_all('Licence', filters={'enabled': 1}, fields=['name'])
    for lic in licences:
        licence = frappe.get_doc('Licence', lic['name'])
        customer = frappe.get_doc('Customer', licence.customer)
        if customer:
            if customer.steuerregion == "AT":
                licence.taxes_and_charges = "Verkaufssteuern Inland 20p (022) - FZAT"
            else:
                licence.taxes_and_charges = "Verkaufssteuern Leistungen EU,DRL (021) - FZAT"
            licence.save()
            print("{0}: {2} updated ({1})".format(customer.name, customer.steuerregion, licence.name))
        else:
            print("{0} has no customer".format(licence.name))
    return

# import HS codes
def import_hs_codes(filename):
    # open workbook
    print("Loading {0}...".format(filename))
    workbook = load_workbook(filename)
    worksheet = workbook["Data Import Template"]
    for row in worksheet.iter_rows('A{}:E{}'.format(21, worksheet.max_row)):
        # loop through all customers
        cells = row
        print("cells: {0}".format(len(cells)))
        if len(cells) >= 4:
            # check if item exists
            matches_by_id = frappe.get_all("Item", filters={'name': cells[2].value}, fields=['name'])
            print("Item: {0}".format(cells[2].value))
            if matches_by_id:
                # found item, update
                print("updating... ({0}, {1}, {2}".format(matches_by_id[0]['name'],cells[3].value,cells[4].value))
                item = frappe.get_doc("Item", matches_by_id[0]['name'])
                item.country_of_origin = unicode(cells[3].value)
                item.customs_tariff_number = unicode(cells[4].value)
                item.save()
    print("List completed")
    return

def enable_vr():
    doctypes = frappe.get_all("DocType", fields=['name'])
    for doctype in doctypes:
        print("Adding {0}...".format(doctype['name']))
        new_access = frappe.get_doc({
          'doctype': 'Custom DocPerm',
          'parenttype': 'DocType',
          'parentfield': 'permissions',
          'parent': doctype['name'],
          'role': "Verwaltungsrat",
          'create': 0,
          'share': 0,
          'export': 0,
          'cancel': 0,
          'submit': 0,
          'write': 0,
          'print': 1,
          'import': 0,
          'read': 1,
          'report': 1,
          'set_user_permission': 0,
          'amend': 0,
          'email': 1,
          'delete': 0
        })
        new_access.insert()
    return

def correct_dl_kst():
    gl_entries = frappe.get_all("GL Entry", filters=[['voucher_no', 'LIKE', 'LS-%']], fields=['name',  'voucher_no'])
    for gl_entry in gl_entries:
        if gl_entry['voucher_no']:
            dl = frappe.get_doc("Delivery Note", gl_entry['voucher_no'])
            customer = frappe.get_doc("Customer", dl.customer)
            #gl_record = frappe.get_doc("GL Entry", gl_entry['name'])
            print("Updating {0} ({3}) from {1} to {2}".format(gl_entry['name'], customer.name, customer.kostenstelle, gl_entry['voucher_no']))
            #gl_record.cost_center = customer.kostenstelle
            #gl_record.save(ignore_permissions=True)
            sql_query = """UPDATE `tabGL Entry` SET `cost_center` = "{kst}" WHERE `name` = "{name}";""".format(name=gl_entry['name'], kst=customer.kostenstelle)
            frappe.db.sql(sql_query, as_dict=True)
    return

def send_customers_to_zsw(tenant):
    print("Sending all customers to ZSW...")
    customers = frappe.get_all("Customer", fields=['name'])
    for customer_nummber in customers:
        customer = frappe.get_doc("Customer", customer_number)
        if tenant.lower() == "ch":
            zsw_ref = "CH" + customer_number[2:]
        else:
            zsw_ref = customer_number[2:]
        if customer.disabled:
            active = False
        else:
            active = True
        print("Updating {0} > {1} (active: {2}, kst: {2}".format(customer_number, zsw_ref, active, customer.kostenstelle))
        create_update_customer(zsw_ref, customer.customer_name, active=active, kst=customer.kostenstelle)
    print("Done. Please check the ERP error log for potential errors.")
    return
